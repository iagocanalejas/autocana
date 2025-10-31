import argparse
import calendar
import logging
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

from openpyxl.drawing.image import Image
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

import autocana.constants as C
from autocana.data.config import load_user_config
from autocana.data.private import PrivateConfig

logger = logging.getLogger("autocana")


@dataclass
class TSHConfig:
    private: PrivateConfig

    # only from 'config.yaml#invoicing'
    activity_id: str
    contract_number: str
    customer_contract: int
    extension_number: int

    rest_days: list[int] = field(default_factory=list)
    _month: int | None = None
    _output_file: str | None = None
    _output_dir: Path | None = None

    @property
    def month(self) -> int:
        return self._month if self._month is not None else datetime.now(timezone.utc).month

    @property
    def file_name(self) -> str:
        if self._output_file:
            return self._output_file
        today = datetime.now(timezone.utc).replace(month=self.month)
        month = today.replace(day=calendar.monthrange(today.year, today.month)[1])

        name_parts = self.private.full_name.split()
        name = f"{name_parts[1][:6]}{name_parts[0][:2]}".lower()
        return f"TSH_{name}_{month.strftime('%Y%m%d').lower()}.xlsx"

    @property
    def output_path(self) -> str:
        if self._output_dir:
            return f"{self._output_dir}/{self.file_name}"
        return self.file_name

    @classmethod
    def load(cls) -> "TSHConfig":
        yaml_cfg = load_user_config()
        invoicing_cfg = yaml_cfg["invoicing"]
        return cls(
            private=PrivateConfig.load(yaml_cfg["private"]),
            activity_id=invoicing_cfg["activity_id"],
            contract_number=invoicing_cfg["contract_number"],
            customer_contract=invoicing_cfg["customer_contract"],
            extension_number=invoicing_cfg["extension_number"],
        )

    def with_params(self, params: argparse.Namespace) -> "TSHConfig":
        self.rest_days = params.skip
        self._month = params.month
        self._output_file = params.output
        if params.output_dir:
            dir = Path(params.output_dir)
            if not dir.exists() and dir.is_dir():
                raise ValueError(f"No dir found in {params.output_dir}")
            self._output_dir = dir
        return self


def fill_worksheet(config: TSHConfig, ws: Worksheet) -> Worksheet:
    tsh_date = datetime.now(timezone.utc).replace(month=config.month)
    ws["AD4"] = tsh_date.strftime("%B")
    ws["AJ4"] = tsh_date.year
    ws["A10"] = f"{config.activity_id}"
    ws["B10"] = "Cronos INT"
    ws["C10"] = "1"
    ws["D10"] = f"TM - SC: {config.extension_number}"
    ws["E10"] = "BI"
    ws["R37"] = tsh_date.strftime("%d/%m/%Y")
    return ws


def fill_worked_days(config: TSHConfig, ws: Worksheet) -> Worksheet:
    tsh_date = datetime.now(timezone.utc).replace(month=config.month)
    weekday, days_in_month = calendar.monthrange(tsh_date.year, tsh_date.month)
    current_col = column_index_from_string("H")
    for day_number in range(1, days_in_month + 1):
        current_col += 1
        if weekday in (5, 6):  # skip weekends
            weekday = (weekday + 1) % 7
            continue

        row = 9 if day_number in config.rest_days else 10
        ws.cell(row=row, column=current_col, value=8)
        weekday = (weekday + 1) % 7
    return ws


def sign_worksheet_if_configured(ws: Worksheet) -> Worksheet:
    if not C.SIGNATURE_FILE_PATH.is_file():
        logger.error("no signature file found, skipping adding signature.")
        return ws

    img = Image(str(C.SIGNATURE_FILE_PATH))
    img.width = 200
    img.height = 95
    ws.add_image(img, "W33")
    return ws
